import streamlit as st
from pathlib import Path
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import io
from config import IMGS_FOLDER_WSL, JSON_PATH_MEALS
from src.gendiet import get_dishes, get_available_dish_types, get_dish_id_by_name
from src.utils import clean_img_name


# --- LOGIC ---

def macro_bar(kcal, protein, carbs, fat):
    st.markdown(f"""
        <div style="font-size:0.9rem; color:#ccc; padding: 0px 0">
            🔥 <b style="color:white">{kcal}</b> kcal &nbsp;|&nbsp;
            B: <b style="color:#4fc3f7">{protein}g</b> &nbsp;
            W: <b style="color:#81c784">{carbs}g</b> &nbsp;
            T: <b style="color:#ffb74d">{fat}g</b>
        </div>
    """, unsafe_allow_html=True)

def generate_shopping_list(selected_meals, diet_dict):
    """Build a combined shopping list from selected weekly meals."""
    shopping = {}
    all_dishes = {
        d["dish_name"]: d
        for dishes in diet_dict.values()
        for d in dishes
    }
    for meal_name in selected_meals:
        dish = all_dishes.get(meal_name)
        if not dish:
            continue
        for ing in dish["ingredients"]:
            key = ing["name"]
            if key not in shopping:
                shopping[key] = {"qty": 0, "unit": ing["unit"], "type": ing.get("type", "")}
            try:
                shopping[key]["qty"] += float(ing["qty"])
            except (ValueError, TypeError):
                pass
    return shopping


def get_excel_binary(shopping: dict) -> bytes:
    """Generate an Excel shopping list file in memory and return bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Shopping List"

    header_fill = PatternFill("solid", fgColor="D7E4BC")
    header_font = Font(bold=True)
    thin         = Side(style="thin")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)
    center       = Alignment(horizontal="center", vertical="center")
    left_al      = Alignment(horizontal="left",   vertical="center")
    middle       = Alignment(vertical="center")

    headers    = ["✓", "Składnik", "Ilość", "Dział"]
    col_widths = [6, 40, 14, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, (ingredient, details) in enumerate(shopping.items(), start=2):
        try:
            qty_val = round(float(details["qty"]), 2)
            qty_str = str(int(qty_val)) if qty_val == int(qty_val) else str(qty_val)
        except (ValueError, TypeError):
            qty_str = str(details["qty"])

        qty_display = f"{qty_str} {details['unit']}".strip()

        # Column A — empty, designed to add checkboxes manually
        a = ws.cell(row=row_idx, column=1, value=None)
        a.border = border

        b = ws.cell(row=row_idx, column=2, value=ingredient)
        b.alignment = middle
        b.border    = border

        c = ws.cell(row=row_idx, column=3, value=qty_display)
        c.alignment = left_al
        c.border    = border

        d = ws.cell(row=row_idx, column=4, value=details.get("type", ""))
        d.alignment = middle
        d.border    = border

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.dialog("🛒 Twoja Lista Zakupów")
def show_shopping_list_dialog(shopping):
    st.write("Składniki potrzebne na cały tydzień:")
    for name, data in shopping.items():
        st.write(f"👉 **{name}**: {round(data['qty'], 2)} {data['unit']}")
    st.divider()
    excel_data = get_excel_binary(shopping)
    st.download_button(
        label="Pobierz listę Excel 📥",
        data=excel_data,
        file_name="shopping_list.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def get_wsl_image_path(dish_name):
    clean_name = clean_img_name(dish_name)
    return Path(IMGS_FOLDER_WSL) / f"{clean_name}.png"


# --- CONFIG ---

st.set_page_config(layout="wide", page_title="Diet Planner", page_icon="🥗")

diff_map = {1: "Łatwo", 2: "Średnio", 3: "Trudno"}
prep_map = {1: "Szybko", 2: "Średnio", 3: "Długo"}

# Load trainer's macro plan from JSON
with open(JSON_PATH_MEALS, "r", encoding="utf-8") as f:
    trainer_plan = json.load(f)

all_meal_types = [m["meal_type"] for m in trainer_plan]

# Exclude protein shake from menu planner
EXCLUDED_FROM_WEEK = {"Białko"}


# --- HELPERS ---

def fetch_diet(macro_plan: dict, dish_limit: int):
    """
    Fetch dishes from DB for each meal type based on current macro targets.
    Returns a dict {meal_type: [dish, ...]} and a list of meal types with no results.
    """
    available_types = get_available_dish_types()
    result   = {}
    warnings = []

    for meal_type in available_types:
        targets = macro_plan.get(meal_type)
        if not targets:
            continue
        dishes = get_dishes(
            dish_type=meal_type,
            kcal_target=targets["kcal"],
            protein_target=targets["proteins"],
            carbs_target=targets["carbs"],
            fat_target=targets["fats"],
            limit=dish_limit,
        )
        if dishes:
            result[meal_type] = dishes
        else:
            warnings.append(meal_type)

    return result, warnings


# --- SESSION STATE ---

if "macro_plan" not in st.session_state:
    st.session_state.macro_plan = {
        m["meal_type"]: {
            "kcal":     int(m["meal_kcal"]),
            "proteins": int(m["meal_proteins"]),
            "carbs":    int(m["meal_carbs"]),
            "fats":     int(m["meal_fats"]),
        }
        for m in trainer_plan
    }

if "dish_limit" not in st.session_state:
    st.session_state.dish_limit = 8

if "diet_dict" not in st.session_state:
    st.session_state.diet_dict = None  # populated on first render or after Recalculate


# --- MAIN UI ---

st.title("🥗 Diet Planner")

# Macro adjustment panel
with st.expander("⚙️ Dostosuj makro", expanded=False):
    st.caption("Domyślne wartości pochodzą z planu ułożonego przez trenera.")

    edited_plan = {}

    cols_header = st.columns([3, 1, 1, 1, 1])
    cols_header[0].markdown("**Posiłek**")
    cols_header[1].markdown("**kcal**")
    cols_header[2].markdown("**Białko (g)**")
    cols_header[3].markdown("**Węglowodany (g)**")
    cols_header[4].markdown("**Tłuszcze (g)**")

    for meal_type in all_meal_types:
        current = st.session_state.macro_plan[meal_type]
        cols = st.columns([3, 1, 1, 1, 1])
        cols[0].markdown(f"**{meal_type}**")
        kcal     = cols[1].number_input("kcal",    value=current["kcal"],     min_value=0, step=10, key=f"edit_kcal_{meal_type}",  label_visibility="collapsed")
        proteins = cols[2].number_input("protein", value=current["proteins"], min_value=0, step=1,  key=f"edit_prot_{meal_type}",  label_visibility="collapsed")
        carbs    = cols[3].number_input("carbs",   value=current["carbs"],    min_value=0, step=1,  key=f"edit_carbs_{meal_type}", label_visibility="collapsed")
        fats     = cols[4].number_input("fat",     value=current["fats"],     min_value=0, step=1,  key=f"edit_fats_{meal_type}",  label_visibility="collapsed")
        edited_plan[meal_type] = {"kcal": kcal, "proteins": proteins, "carbs": carbs, "fats": fats}

    col_limit, col_btn = st.columns([1, 3])
    dish_limit = col_limit.number_input(
        "Maks. dań na posiłek",
        min_value=1, max_value=50,
        value=st.session_state.dish_limit,
        step=1,
        key="edit_limit"
    )
    if st.button("🔄 Przelicz", type="primary", width='content'):
        st.session_state.macro_plan = edited_plan
        st.session_state.dish_limit = dish_limit
        st.session_state.diet_dict  = None

# Fetch dishes on first load or after Recalculate
if st.session_state.diet_dict is None:
    with st.spinner("Pobieranie dań z bazy danych..."):
        diet_dict, missing = fetch_diet(
            st.session_state.macro_plan,
            st.session_state.dish_limit
        )
    st.session_state.diet_dict = diet_dict
    if missing:
        st.info(f"Brak dań w bazie danych dla: **{', '.join(missing)}**.")
else:
    diet_dict = st.session_state.diet_dict

# Dish browser
col_radio, col_sort, col_order = st.columns([3, 2, 1])

with col_radio:
    meal_type_view = st.radio("Przeglądaj według typu posiłku:", list(diet_dict.keys()), horizontal=True)

with col_sort:
    sort_by = st.selectbox("Sortuj:", [
        "Domyślnie", "Kalorie", "Białko", "Węglowodany",
        "Tłuszcze", "Czas przygotowania", "Poziom trudności", "Liczba składników",
    ], label_visibility="collapsed")

with col_order:
    descending = st.toggle("↓ descending", value=False)

sort_key_map = {
    "Domyślnie":            lambda d: 0,
    "Kalorie":           lambda d: d["kcal"],
    "Białko":            lambda d: d["protein_g"],
    "Węglowodany":              lambda d: d["carbs_g"],
    "Tłuszcze":                lambda d: d["fat_g"],
    "Czas przygotowania":          lambda d: int(d["prep_time"]),
    "Poziom trudności":         lambda d: int(d["difficulty"]),
    "Liczba składników":  lambda d: len(d["ingredients"]),
}

dishes_to_show = diet_dict[meal_type_view]
if sort_by != "Default":
    dishes_to_show = sorted(dishes_to_show, key=sort_key_map[sort_by], reverse=descending)

for dish in dishes_to_show:
    p_time  = prep_map.get(int(dish["prep_time"]),  "?")
    d_level = diff_map.get(int(dish["difficulty"]), "?")
    n_ing   = len(dish["ingredients"])
    label = (
        f"{dish['dish_name']}   |   "
        f"🔥 {dish['kcal']} kcal   "
        f"B: {dish['protein_g']}g  W: {dish['carbs_g']}g  T: {dish['fat_g']}g   |   "
        f"🕒 {p_time}  💪 {d_level}  🛒 {n_ing} skł."
    )
    with st.expander(label):
        col1, col2 = st.columns([1, 2])
        with col1:
            wsl_path = get_wsl_image_path(dish["dish_name"])
            if wsl_path.exists():
                st.image(str(wsl_path))
            else:
                st.info("Brak zdjęcia")
            st.markdown("### 👨‍🍳 Przepis")
            st.write(dish["recipe_text"])
        with col2:
            st.markdown(f"**Makro:** B: {dish['protein_g']}g | W: {dish['carbs_g']}g | T: {dish['fat_g']}g")
            st.write(f"🕒 {p_time} | 💪 {d_level} | 🛒 {n_ing} ingredients")
            st.markdown("### 🛒 Składniki")
            for ing in dish["ingredients"]:
                st.write(f"• {ing['name']} - {ing['qty']} {ing['unit']}")


# --- SIDEBAR - WEEKLY MENU ---

st.sidebar.title("📅 Tygodniowe menu")

days = ["Poniedziałek", "Wtorek", "Środa", "Czwartek", "Piątek", "Sobota", "Niedziela"]
meal_types_to_plan = [m for m in diet_dict.keys() if m not in EXCLUDED_FROM_WEEK]

selected_meals_week = []

name_to_macros = {
    dish['dish_name']: dish
    for m_type in meal_types_to_plan
    for dish in diet_dict[m_type]
}

for day in days:
    with st.sidebar.expander(day):
        selected_meals_day = []
        for m_type in meal_types_to_plan:
            options = [d["dish_name"] for d in diet_dict[m_type]]
            choice  = st.selectbox(f"{m_type}", options, key=f"{day}_{m_type}")
            selected_meals_week.append(choice)
            selected_meals_day.append(choice)
        day_totals = {
            key: sum(name_to_macros[name][key] for name in selected_meals_day if name in name_to_macros)
            for key in ['kcal', 'protein_g', 'carbs_g', 'fat_g']
        }
        macro_bar(
            day_totals['kcal'],
            day_totals['protein_g'],
            day_totals['carbs_g'],
            day_totals['fat_g']
        )
with st.sidebar:
    week_totals = {
        key: sum(name_to_macros[name][key] for name in selected_meals_week if name in name_to_macros)
        for key in ['kcal', 'protein_g', 'carbs_g', 'fat_g']
    }
    week_avg = {key: round(val / len(days)) for key, val in week_totals.items()}
    st.subheader("Średnia dzienna")
    macro_bar(
        week_avg['kcal'],
        week_avg['protein_g'],
        week_avg['carbs_g'],
        week_avg['fat_g']
    )
st.sidebar.divider()
if st.sidebar.button("🚀 Wygeneruj listę zakupów"):
    shopping_result = generate_shopping_list(selected_meals_week, diet_dict)
    show_shopping_list_dialog(shopping_result)
